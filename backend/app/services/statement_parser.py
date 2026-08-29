"""Universal Card Statement & Transaction History Parser.

Supports:
- Excel files (.xlsx, including non-standard spreadsheet XMLs)
- Excel binary (.xls, including HTML tables saved with .xls extension)
- CSV files (UTF-8, UTF-8-BOM, CP949, EUC-KR)
- PDF statements (KB국민, NH농협, 현대, etc.) with optional password decryption
- Images / Screen captures via Gemini Vision
"""

from __future__ import annotations

import csv
import io
import logging
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any
from uuid import UUID

import openpyxl
import xlrd
from pypdf import PdfReader

from app.models import Category
from app.models.category import TransactionType
from app.services.ai import _extract_json, _gemini_client, is_ai_enabled

logger = logging.getLogger(__name__)


@dataclass
class ParsedStatementItem:
    transaction_date: str  # YYYY-MM-DD
    payee: str
    amount: float
    type: str = "expense"  # 'expense' or 'income'
    currency: str = "KRW"
    memo: str | None = None
    suggested_category_id: UUID | None = None
    suggested_category_name: str | None = None
    card_name: str | None = None
    approval_no: str | None = None


@dataclass
class StatementParseResult:
    card_company: str | None
    total_count: int
    total_amount: float
    items: list[ParsedStatementItem] = field(default_factory=list)
    requires_password: bool = False
    error_message: str | None = None


def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _clean_amount(val: Any) -> float | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("원", "").replace("KRW", "").replace("$", "").replace(" ", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _normalize_date(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None

    # Match YYYY-MM-DD or YYYY/MM/DD or YYYY.MM.DD
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"

    # Match YY.MM.DD or YY/MM/DD or YY-MM-DD (e.g. 26.07.01)
    m = re.match(r"^(\d{2})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        full_y = 2000 + int(y)
        return f"{full_y:04d}-{int(mo):02d}-{int(d):02d}"

    # Match YYYY년 MM월 DD일
    m = re.match(r"^(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"

    # Match YYYYMMDD
    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{mo}-{d}"

    return None


def _detect_card_company(text_corpus: str) -> str | None:
    text = text_corpus.upper()
    if "KB" in text or "국민카드" in text or "KB국민" in text or "포인트리" in text:
        return "KB국민카드"
    if "NH" in text or "농협카드" in text or "NH농협" in text or "채움" in text:
        return "NH농협카드"
    if "현대카드" in text or "HYUNDAI" in text or "M포인트" in text or "SMILECARD" in text:
        return "현대카드"
    if "신한카드" in text or "SHINHAN" in text or "MYSHINHAN" in text:
        return "신한카드"
    if "삼성카드" in text or "SAMSUNG" in text:
        return "삼성카드"
    if "롯데카드" in text or "LOTTE" in text:
        return "롯데카드"
    if "우리카드" in text or "WOORI" in text:
        return "우리카드"
    if "하나카드" in text or "HANA" in text:
        return "하나카드"
    if "BC카드" in text or "비씨카드" in text:
        return "BC카드"
    return None


def _find_header_and_map(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    """Find the table header row index and map column indices for date, payee, amount, etc."""
    header_keywords = {
        "date": ["이용일", "이용일자", "거래일자", "승인일자", "승인일시", "일자", "날짜", "date"],
        "payee": ["이용하신곳", "이용가맹점", "가맹점명", "가맹점", "사용처", "적요", "내용", "상호", "merchant", "payee"],
        "amount": ["국내이용금액", "청구금액", "이용금액", "승인금액", "결제원금", "청구원금", "원금", "금액", "amount", "당월청구금액"],
        "card": ["이용카드명", "이용카드", "카드명", "카드번호", "카드"],
        "method": ["결제방법", "이용구분", "할부/회차", "할부", "할부개월"],
        "status": ["상태", "접수여부", "매입상태", "구분"],
        "approval": ["승인번호"],
    }

    best_row_idx = -1
    best_mapping: dict[str, int] = {}
    best_score = 0

    for r_idx, row in enumerate(rows[:30]):
        row_strs = [_clean_str(c).replace("\n", "").replace(" ", "").lower() for c in row]
        mapping: dict[str, int] = {}
        score = 0

        for col_name, keywords in header_keywords.items():
            for c_idx, cell in enumerate(row_strs):
                if any(kw.lower() in cell for kw in keywords):
                    if col_name not in mapping:
                        mapping[col_name] = c_idx
                        score += 1
                        break

        if "date" in mapping and "payee" in mapping and "amount" in mapping:
            if score > best_score:
                best_score = score
                best_row_idx = r_idx
                best_mapping = mapping

    if best_row_idx >= 0:
        return best_row_idx, best_mapping
    return None


def _is_summary_row(row_strs: list[str]) -> bool:
    joined = "".join(row_strs)
    summary_markers = ["소계", "총합계", "총 합계", "합계", "총 건수", "이용 소계", "기본정보", "고객님이 요청하신"]
    return any(marker in joined for marker in summary_markers)


def _guess_category_by_rules(payee: str, categories: list[Category]) -> tuple[UUID | None, str | None]:
    p = payee.lower()
    name_map = {c.name: c.id for c in categories if c.type == TransactionType.EXPENSE}

    cat_rules = [
        (["스타벅스", "이디야", "투썸", "메가mgc", "빽다방", "컴포즈", "커피", "카페", "베이커리", "뚜레쥬르", "파리바게뜨", "좋은아침", "버터비버"], ["카페/간식", "식비", "간식"]),
        (["식당", "음식점", "식사", "순대국", "코다리", "김밥", "충무김밥", "써브웨이", "치킨", "버거", "맥도날드", "피자", "배달의민족", "요기요", "쿠팡이츠", "풀무원", "맛사랑", "단정"], ["식비", "외식"]),
        (["마트", "이마트", "홈플러스", "롯데마트", "지에스25", "gs25", "씨유", "cu", "세븐일레븐", "이마트24", "농협", "하나로마트", "우체국쇼핑", "g마켓", "옥션", "쿠팡", "네이버페이", "11번가", "aliexpress"], ["마트/생필품", "쇼핑", "식비", "생활비"]),
        (["택시", "카카오t", "카카오_택시", "카카오_개인택시", "그린카", "쏘카", "지하철", "버스", "코레일", "하이패스", "주유소", "gs칼텍스", "sk에너지", "s-oil", "현대오일"], ["교통", "차량/교통", "교통/차량"]),
        (["약국", "의원", "병원", "내과", "외과", "정형외과", "치과", "한의원", "이레미즈", "새희망내과", "대학약국"], ["의료/건강", "의료", "건강"]),
        (["통신", "lgu+", "lguplus", "sk텔레콤", "kt", "통신요금", "알뜰폰"], ["통신/주거", "통신비", "통신"]),
        (["넷플릭스", "유튜브", "쿠팡플레이", "티빙", "왓챠", "디즈니", "cgv", "롯데시네마", "메가박스", "anthropic", "openai", "apple", "구글", "모노맨션", "지방세"], ["문화/여가", "구독", "공과금", "문화/생활"]),
    ]

    for keywords, target_cats in cat_rules:
        if any(kw in p for kw in keywords):
            for t_cat in target_cats:
                for c_name, c_id in name_map.items():
                    if t_cat in c_name or c_name in t_cat:
                        return c_id, c_name

    for c_name, c_id in name_map.items():
        if c_name in p:
            return c_id, c_name

    return None, None


def _parse_raw_table(rows: list[list[Any]], categories: list[Category]) -> StatementParseResult:
    text_corpus = " ".join(" ".join(_clean_str(c) for c in r) for r in rows[:15])
    card_company = _detect_card_company(text_corpus)

    matched = _find_header_and_map(rows)
    if not matched:
        return StatementParseResult(card_company=card_company, total_count=0, total_amount=0.0, items=[])

    header_idx, mapping = matched
    date_col = mapping["date"]
    payee_col = mapping["payee"]
    amount_col = mapping["amount"]
    card_col = mapping.get("card")
    method_col = mapping.get("method")
    status_col = mapping.get("status")
    approval_col = mapping.get("approval")

    items: list[ParsedStatementItem] = []
    total_amount = 0.0

    for r in rows[header_idx + 1:]:
        row_strs = [_clean_str(c) for c in r]
        if not any(row_strs) or _is_summary_row(row_strs):
            continue

        raw_date = r[date_col] if date_col < len(r) else None
        norm_date = _normalize_date(raw_date)
        if not norm_date:
            continue

        raw_payee = _clean_str(r[payee_col]) if payee_col < len(r) else ""
        if not raw_payee or raw_payee == "-":
            continue

        raw_amount = r[amount_col] if amount_col < len(r) else None
        amount = _clean_amount(raw_amount)
        if amount is None or amount == 0:
            continue

        raw_status = _clean_str(r[status_col]) if status_col is not None and status_col < len(r) else ""
        is_cancel = "취소" in raw_status or "취소" in raw_payee or amount < 0

        final_amount = abs(amount)
        txn_type = "income" if is_cancel else "expense"

        memos = []
        if method_col is not None and method_col < len(r):
            m = _clean_str(r[method_col])
            if m and m != "/":
                memos.append(m)
        card_name = None
        if card_col is not None and card_col < len(r):
            c_name = _clean_str(r[card_col])
            if c_name:
                card_name = c_name
                memos.append(c_name)
        if is_cancel:
            memos.append("승인취소")

        approval_no = None
        if approval_col is not None and approval_col < len(r):
            app_no = _clean_str(r[approval_col])
            if app_no:
                approval_no = app_no

        memo_str = " | ".join(memos) if memos else None

        cat_id, cat_name = _guess_category_by_rules(raw_payee, categories)

        items.append(
            ParsedStatementItem(
                transaction_date=norm_date,
                payee=raw_payee,
                amount=final_amount,
                type=txn_type,
                currency="KRW",
                memo=memo_str,
                suggested_category_id=cat_id,
                suggested_category_name=cat_name,
                card_name=card_name,
                approval_no=approval_no,
            )
        )
        total_amount += final_amount

    return StatementParseResult(
        card_company=card_company,
        total_count=len(items),
        total_amount=round(total_amount, 2),
        items=items,
    )


def _parse_xlsx_openpyxl(content: bytes) -> list[list[Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    rows: list[list[Any]] = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        if ws.max_row > 1:
            for r in range(1, ws.max_row + 1):
                row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                rows.append(row_vals)
            break
    return rows


def _parse_xlsx_xml(content: bytes) -> list[list[Any]]:
    """Fallback XML parser for non-standard or slightly corrupt xlsx files."""
    import zipfile
    with zipfile.ZipFile(io.BytesIO(content), "r") as z:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            tree = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for elem in tree.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                shared_strings.append(elem.text or "")

        sheet_files = [f for f in z.namelist() if f.startswith("xl/worksheets/sheet") and f.endswith(".xml")]
        sheet_files.sort()
        if not sheet_files:
            return []

        tree = ET.fromstring(z.read(sheet_files[0]))
        rows: list[list[Any]] = []
        for row_elem in tree.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row"):
            cells = []
            for c in row_elem.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
                cell_type = c.attrib.get("t")
                val_elem = c.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
                val = val_elem.text if val_elem is not None else None
                if cell_type == "s" and val is not None:
                    try:
                        val = shared_strings[int(val)] if int(val) < len(shared_strings) else val
                    except ValueError:
                        pass
                cells.append(val)
            rows.append(cells)
        return rows


def _parse_xls(content: bytes) -> list[list[Any]]:
    head = content[:1000].lower()
    if b"<html" in head or b"<table" in head or b"<!doctype" in head:
        text = content.decode("utf-8", errors="replace")
        trs = re.findall(r"<tr[^>]*>(.*?)</tr>", text, re.DOTALL | re.IGNORECASE)
        rows: list[list[Any]] = []
        for tr in trs:
            # Match both <td> and <th>, taking into account spaces before closing bracket like </td   >
            tds = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]\s*>", tr, re.DOTALL | re.IGNORECASE)
            cleaned = [re.sub(r"<[^>]+>", "", td).strip() for td in tds]
            if any(cleaned):
                rows.append(cleaned)
        return rows
    else:
        wb = xlrd.open_workbook(file_contents=content)
        sheet = wb.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            rows.append(sheet.row_values(r))
        return rows


def _parse_csv(content: bytes) -> list[list[Any]]:
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        try:
            text = content.decode(enc)
            reader = csv.reader(io.StringIO(text))
            return list(reader)
        except UnicodeDecodeError:
            continue
    return []


def _parse_pdf(content: bytes, password: str | None = None) -> tuple[str, bool]:
    """Extract text from PDF, handling encryption/passwords."""
    try:
        reader = PdfReader(io.BytesIO(content))
        if reader.is_encrypted:
            try:
                decrypted = reader.decrypt(password or "")
                if not decrypted:
                    return "", True
            except Exception:
                return "", True

        pages_text = []
        for page in reader.pages:
            txt = page.extract_text()
            if txt:
                pages_text.append(txt)
        return "\n".join(pages_text), False
    except Exception as exc:
        logger.warning(f"pypdf error: {exc}")
        return "", False


def _parse_with_gemini_ai(
    content: bytes,
    media_type: str,
    categories: list[Category],
) -> StatementParseResult:
    """Use Gemini Multimodal Vision / Document AI for complex PDFs or images."""
    if not is_ai_enabled():
        raise RuntimeError("AI parsing is disabled (no GEMINI_API_KEY)")

    expense_cats = [c.name for c in categories if c.type == TransactionType.EXPENSE]
    catalog = ", ".join(expense_cats) or "(none)"

    prompt = (
        "You are an expert financial assistant that reads Korean credit card statements (신용카드 이용대금명세서 / 이용내역서). "
        "Extract all individual card transactions into structured JSON. "
        "Always reply with JSON only.\n\n"
        f"Available expense categories: {catalog}\n\n"
        "Return this exact JSON schema: "
        '{\n'
        '  "card_company": "<e.g. KB국민카드, NH농협카드, 현대카드 or null>",\n'
        '  "items": [\n'
        '    {\n'
        '      "transaction_date": "YYYY-MM-DD",\n'
        '      "payee": "<merchant/store name>",\n'
        '      "amount": <number, strictly positive>,\n'
        '      "type": "<expense or income (income if cancel/refund)>",\n'
        '      "currency": "KRW",\n'
        '      "memo": "<card name, installment, or brief note or null>",\n'
        '      "suggested_category_name": "<exact name from available categories or null>"\n'
        '    }\n'
        '  ]\n'
        '}'
    )

    client = _gemini_client()
    from app.core.config import settings
    from google.genai import types as genai_types

    response = client.models.generate_content(
        model=settings.GEMINI_MODEL,
        contents=[
            genai_types.Part.from_bytes(data=content, mime_type=media_type),
            prompt,
        ],
        config=genai_types.GenerateContentConfig(
            system_instruction="Extract all card statement transactions accurately into JSON.",
            response_mime_type="application/json",
            temperature=0.1,
        ),
    )
    parsed = _extract_json(response.text or "{}")

    card_company = parsed.get("card_company")
    raw_items = parsed.get("items") or []
    items: list[ParsedStatementItem] = []
    total_amount = 0.0

    name_to_id = {c.name: c.id for c in categories}

    for it in raw_items:
        if not isinstance(it, dict):
            continue
        p_name = _clean_str(it.get("payee"))
        p_date = _normalize_date(it.get("transaction_date"))
        p_amt = _clean_amount(it.get("amount"))
        if not p_name or not p_date or not p_amt:
            continue

        p_type = str(it.get("type") or "expense").lower()
        cat_name = it.get("suggested_category_name")
        cat_id = name_to_id.get(cat_name) if cat_name else None

        items.append(
            ParsedStatementItem(
                transaction_date=p_date,
                payee=p_name,
                amount=p_amt,
                type=p_type,
                currency=str(it.get("currency") or "KRW"),
                memo=it.get("memo"),
                suggested_category_id=cat_id,
                suggested_category_name=cat_name,
            )
        )
        total_amount += p_amt

    return StatementParseResult(
        card_company=card_company,
        total_count=len(items),
        total_amount=round(total_amount, 2),
        items=items,
    )


def parse_statement_file(
    content: bytes,
    filename: str,
    content_type: str | None,
    categories: list[Category],
    password: str | None = None,
) -> StatementParseResult:
    """Master entry point for parsing any credit card statement file."""
    fname = filename.lower()

    # 1. Image files -> Gemini Vision
    if (content_type and content_type.startswith("image/")) or fname.endswith((".jpg", ".jpeg", ".png", ".webp")):
        mime = content_type if (content_type and content_type.startswith("image/")) else "image/jpeg"
        return _parse_with_gemini_ai(content, mime, categories)

    # 2. PDF files
    if fname.endswith(".pdf") or content_type == "application/pdf":
        pdf_text, is_encrypted = _parse_pdf(content, password)
        if is_encrypted:
            return StatementParseResult(
                card_company=None,
                total_count=0,
                total_amount=0.0,
                requires_password=True,
                error_message="비밀번호로 보호된 PDF 명세서입니다. 비밀번호(생년월일 6자리 등)를 입력해주세요.",
            )
        if is_ai_enabled():
            try:
                return _parse_with_gemini_ai(content, "application/pdf", categories)
            except Exception as e:
                logger.warning(f"AI PDF parse error: {e}")

    # 3. XLSX files
    if fname.endswith(".xlsx"):
        rows: list[list[Any]] = []
        try:
            rows = _parse_xlsx_openpyxl(content)
        except Exception:
            rows = _parse_xlsx_xml(content)
        if rows:
            res = _parse_raw_table(rows, categories)
            if res.total_count > 0:
                return res

    # 4. XLS files
    if fname.endswith(".xls"):
        rows = _parse_xls(content)
        if rows:
            res = _parse_raw_table(rows, categories)
            if res.total_count > 0:
                return res

    # 5. CSV files
    if fname.endswith(".csv") or content_type == "text/csv":
        rows = _parse_csv(content)
        if rows:
            res = _parse_raw_table(rows, categories)
            if res.total_count > 0:
                return res

    # 6. Fallback to AI
    if is_ai_enabled():
        mime = content_type or "application/octet-stream"
        if fname.endswith(".pdf"):
            mime = "application/pdf"
        elif fname.endswith((".xlsx", ".xls")):
            mime = "text/plain"
        try:
            return _parse_with_gemini_ai(content, mime, categories)
        except Exception as e:
            logger.warning(f"AI parsing fallback failed: {e}")

    return StatementParseResult(card_company=None, total_count=0, total_amount=0.0, items=[])
