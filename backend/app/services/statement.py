"""Parse Korean card/bank statement spreadsheets (.xls / .xlsx) into ledger rows.

Card companies export slightly different layouts, so instead of hard-coding one
format we auto-detect the header row (the row that has a date-like, a
merchant-like and an amount-like column) and map columns by keyword.

Rules:
- Domestic amount (국내...원) -> expense in KRW.
- Foreign amount (해외...$) with no domestic amount -> expense in that currency.
- Cancelled rows (상태 contains 취소) -> registered as INCOME (a refund), so they
  offset the original charge and keep the balance correct.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

DATE_HEADERS = ["이용일", "거래일", "승인일", "이용일자", "거래일자", "매출일", "사용일", "날짜"]
PAYEE_HEADERS = ["이용하신곳", "가맹점", "사용처", "이용가맹점", "가맹점명", "이용장소", "적요", "내용"]
AMOUNT_HEADERS = ["국내이용금액", "이용금액", "승인금액", "결제금액", "사용금액", "금액"]
STATUS_HEADERS = ["상태", "취소", "승인구분"]
AMOUNT_EXCLUDE = ["해외", "$", "할인", "적립", "포인트", "예정", "잔액", "수수료"]


@dataclass
class StatementRow:
    transaction_date: date
    amount: Decimal
    currency: str
    txn_type: str  # "expense" | "income"
    payee: str | None
    memo: str | None


@dataclass
class StatementParseResult:
    rows: list[StatementRow] = field(default_factory=list)
    income_from_cancel: int = 0  # cancelled rows registered as income
    foreign: int = 0  # rows registered in a non-KRW currency
    skipped: int = 0  # rows with no usable amount
    errors: list[str] = field(default_factory=list)


def _norm(value: object) -> str:
    return re.sub(r"\s+", "", "" if value is None else str(value))


def _load_grid(filename: str | None, data: bytes) -> list[list[object]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _grid_xlsx(data)
    if name.endswith(".xls"):
        return _grid_xls(data)
    try:
        return _grid_xls(data)
    except Exception:
        return _grid_xlsx(data)


def _grid_xls(data: bytes) -> list[list[object]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    grid: list[list[object]] = []
    for r in range(sheet.nrows):
        row: list[object] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, book.datemode).date())
            else:
                row.append(cell.value)
        grid.append(row)
    return grid


def _grid_xlsx(data: bytes) -> list[list[object]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _find_header(grid: list[list[object]]) -> tuple[int, list[str]]:
    for i, row in enumerate(grid):
        cells = [_norm(x) for x in row]
        has_date = any(any(k in c for k in DATE_HEADERS) for c in cells)
        has_amount = any(any(k in c for k in AMOUNT_HEADERS) for c in cells)
        has_payee = any(any(k in c for k in PAYEE_HEADERS) for c in cells)
        if has_date and has_amount and has_payee:
            return i, cells
    raise ValueError("헤더 행을 찾을 수 없습니다 (이용일/가맹점/금액 열 필요)")


def _pick(cells: list[str], keys: list[str], exclude: tuple[str, ...] = ()) -> int | None:
    best_score, best_idx = 0, None
    for i, c in enumerate(cells):
        if not any(k in c for k in keys):
            continue
        score = sum(2 for k in keys if k in c)
        if "원" in c:
            score += 1
        if any(b in c for b in exclude):
            score -= 5
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def _find_foreign(cells: list[str]) -> tuple[int | None, str]:
    """Return (column index of a foreign-amount column, its currency code)."""
    for i, c in enumerate(cells):
        if "해외" in c and ("금액" in c or "$" in c or "USD" in c.upper()):
            return i, _currency_from_header(c)
    return None, "USD"


def _currency_from_header(header: str) -> str:
    h = header.upper()
    if "$" in header or "USD" in h or "달러" in header:
        return "USD"
    if "€" in header or "EUR" in h or "유로" in header:
        return "EUR"
    if "¥" in header or "JPY" in h or "엔" in header:
        return "JPY"
    return "USD"


def _to_amount(value: object) -> Decimal:
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    cleaned = re.sub(r"[^\d.-]", "", str(value or ""))
    if not cleaned:
        return Decimal(0)
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal(0)


def _to_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10] if "-" in fmt else s[:8], fmt).date()
        except ValueError:
            continue
    raise ValueError(f"날짜 형식 인식 불가: {s}")


def parse_statement(filename: str | None, data: bytes) -> StatementParseResult:
    grid = _load_grid(filename, data)
    if not grid:
        raise ValueError("빈 파일입니다")

    header_idx, cells = _find_header(grid)
    date_col = _pick(cells, DATE_HEADERS)
    payee_col = _pick(cells, PAYEE_HEADERS, exclude=("카드", "고객"))
    amount_col = _pick(cells, AMOUNT_HEADERS, exclude=tuple(AMOUNT_EXCLUDE))
    status_col = _pick(cells, STATUS_HEADERS)
    foreign_col, foreign_currency = _find_foreign(cells)
    if date_col is None or (amount_col is None and foreign_col is None):
        raise ValueError("날짜/금액 열을 인식하지 못했습니다")

    result = StatementParseResult()
    for line_no, row in enumerate(grid[header_idx + 1 :], start=header_idx + 2):
        raw_date = row[date_col] if date_col < len(row) else None
        if raw_date is None or (isinstance(raw_date, str) and not raw_date.strip()):
            continue
        try:
            is_cancel = False
            if status_col is not None and status_col < len(row):
                is_cancel = "취소" in _norm(row[status_col])

            krw = _to_amount(row[amount_col]) if amount_col is not None and amount_col < len(row) else Decimal(0)
            fx = _to_amount(row[foreign_col]) if foreign_col is not None and foreign_col < len(row) else Decimal(0)

            if krw > 0:
                amount, currency, is_foreign = krw, "KRW", False
            elif fx > 0:
                amount, currency, is_foreign = fx, foreign_currency, True
            else:
                result.skipped += 1
                continue

            txn_date = _to_date(raw_date)
            payee = None
            if payee_col is not None and payee_col < len(row) and row[payee_col] not in (None, ""):
                payee = str(row[payee_col]).strip()[:100]

            result.rows.append(
                StatementRow(
                    transaction_date=txn_date,
                    amount=amount,
                    currency=currency,
                    txn_type="income" if is_cancel else "expense",
                    payee=payee,
                    memo="카드 취소/환불" if is_cancel else None,
                )
            )
            if is_cancel:
                result.income_from_cancel += 1
            if is_foreign:
                result.foreign += 1
        except (ValueError, InvalidOperation, TypeError) as exc:
            result.errors.append(f"{line_no}행: {exc}")

    return result
